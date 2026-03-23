#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
네이버 카페 크롤링 스크립트
키워드 기반 게시글/댓글 모니터링
"""

import json
import logging
import multiprocessing
import os
import random
import re
import sys
import time
import traceback
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from urllib.parse import quote

from playwright.sync_api import sync_playwright, Page, Browser, Frame, TimeoutError as PlaywrightTimeoutError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from pydantic import BaseModel, Field, field_validator
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

# Windows 콘솔 인코딩 설정
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')


# 타입 별칭
FrameLike = Union[Page, Frame]


@dataclass(frozen=True)
class Timeouts:
    """타임아웃 설정"""
    PAGE_LOAD = 30000  # 30초
    ELEMENT_WAIT = 5000  # 5초
    COMMENT_LOAD = 3000  # 3초
    NETWORK_IDLE = 10000  # 10초


@dataclass(frozen=True)
class WaitTimes:
    """대기 시간 설정 (밀리초)"""
    AFTER_LOGIN = 2000
    AFTER_PAGE_LOAD = 500
    AFTER_IFRAME_SWITCH = 1000
    AFTER_COMMENT_CLICK = 2000
    BETWEEN_PAGES = 1000
    SCROLL_INTERVAL = 300


@dataclass(frozen=True)
class RetryConfig:
    """재시도 설정"""
    MAX_ATTEMPTS = 3
    MIN_WAIT = 1
    MAX_WAIT = 10


@dataclass(frozen=True)
class CrawlerConstants:
    """크롤러 상수"""
    MAX_SCROLL_ATTEMPTS = 5
    LOGIN_TIMEOUT_SECONDS = 120
    LOGIN_CHECK_INTERVAL = 10


class CafeInfo(BaseModel):
    """카페 정보"""
    cafe_id: str = Field(..., pattern=r'^\d+$', description="카페 ID")
    cafe_name: str = Field(..., min_length=1, description="카페 이름")
    cafe_url: str = Field(..., pattern=r'^https?://', description="카페 URL")


class AccountConfig(BaseModel):
    """계정 정보"""
    naver_id: str = Field(..., min_length=1, description="네이버 아이디")
    naver_password: str = Field(..., min_length=1, description="네이버 비밀번호")
    assigned_cafes: List[str] = Field(..., min_length=1, description="담당 카페 목록")
    group_name: str = Field(..., min_length=1, description="그룹 이름")


class CrawlerSettings(BaseModel):
    """크롤러 설정 (Pydantic 검증)"""
    accounts: List[AccountConfig] = Field(..., min_length=1, description="계정 목록")
    cafes: List[CafeInfo] = Field(..., min_length=1, description="카페 목록")
    keywords: List[str] = Field(..., min_length=1, description="검색 키워드 목록")
    output_folder: str = Field(default='results', description="출력 폴더")
    log_folder: str = Field(default='logs', description="로그 폴더")

    @field_validator('keywords')
    @classmethod
    def validate_keywords(cls, v):
        if not v or len(v) == 0:
            raise ValueError("keywords must have at least one item")
        return v

    @field_validator('cafes')
    @classmethod
    def validate_cafes(cls, v):
        if not v or len(v) == 0:
            raise ValueError("cafes must have at least one item")
        return v


@dataclass(frozen=True)
class Selectors:
    """CSS 선택자 상수"""

    # 로그인
    LOGIN_ID: str = '#id'
    LOGIN_PW: str = '#pw'
    LOGIN_BUTTON: str = '#log\\.login'
    LOGIN_BUTTON_ALT: str = 'button[type="submit"]'
    LOGIN_ERROR: str = '.error_msg, .alert_msg, .err'

    # 검색 결과
    ARTICLE_LINKS: tuple = (
        'div.inner_list a.article',
        'a.article',
        'div.inner_list a',
        'a[href*="/articles/"]',
        '.article-board a.article',
        'a[href*="ArticleRead"]'
    )

    # 게시글 정보
    TITLE: str = 'h3.title_text, .title_text, .title_area'
    AUTHOR: str = 'button.nickname, .nickname, .nick_name, .article_writer .nickname'
    DATE: str = 'span.date, .date, .article_info .date, .write_date'
    CONTENT: str = '.article_viewer, .se-main-container, .content, #content'
    LIKES: tuple = (
        'em._count',
        'em.u_cnt',
        '._count',
        '.u_cnt',
        '.like_count',
        '.like-count'
    )

    # 댓글
    COMMENT_BUTTON: tuple = (
        'a.button_comment',
        'a[class*="button_comment"]',
        'button[class*="comment"]',
        'a[href*="comment"]',
        '*[class*="button_comment"]'
    )
    COMMENT_ITEMS: tuple = (
        'ul.comment_list li.CommentItem',
        'li.CommentItem',
        'ul.comment_list li',
        'div.comment_area'
    )
    COMMENT_AUTHOR: str = 'a.comment_nickname'
    COMMENT_TEXT: str = 'span.text_comment'


class NaverCafeCrawler:
    """네이버 카페 크롤러 클래스"""

    def __init__(self, config_path: str = "config.json", account_info: AccountConfig = None, group_name: str = None):
        """초기화"""
        # 설정 로드 및 검증
        with open(config_path, 'r', encoding='utf-8') as f:
            config_dict = json.load(f)
        self.config = CrawlerSettings(**config_dict)

        self.account_info = account_info
        self.group_name = group_name

        self.logger = self._setup_logger()
        self.collected_data: List[Dict[str, Any]] = []

        # 브라우저 관련
        self.playwright = None
        self.browser: Browser = None
        self.context = None
        self.page: Page = None

        # 상수
        self.selectors = Selectors()
        self.timeouts = Timeouts()
        self.wait_times = WaitTimes()
        self.constants = CrawlerConstants()

        # 타이머 및 재시작 관련
        self.process_start_time = time.time()
        self.last_restart_time = time.time()
        self.restart_interval = 1800  # 30분 (초 단위) - 메모리 최적화를 위해 필요시 1200 (20분) 또는 1500 (25분)으로 조정 가능
        self.restart_count = 0

        # 중복 URL 체크용
        self.existing_urls = set()

    def _setup_logger(self) -> logging.Logger:
        """로거 설정"""
        log_folder = Path(self.config.log_folder)
        log_folder.mkdir(parents=True, exist_ok=True)

        log_filename = log_folder / f"crawler_{self.group_name}_{datetime.now().strftime('%Y%m%d')}.log"

        logger = logging.getLogger(f'NaverCafeCrawler_{self.group_name}')
        logger.setLevel(logging.INFO)

        if logger.handlers:
            logger.handlers.clear()

        # 파일 핸들러 (DEBUG)
        file_handler = logging.FileHandler(log_filename, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)

        # 콘솔 핸들러 (INFO)
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)

        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )

        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)

        logger.addHandler(file_handler)
        logger.addHandler(console_handler)

        return logger

    def _get_week_info(self) -> tuple:
        """파일명용 월/주차 계산 - 이번 주기의 화요일 기준
        (오늘이 화요일이면 오늘, 아니면 가장 가까운 다음 화요일)
        """
        today = datetime.now()
        weekday = today.weekday()  # 0=월, 1=화, ..., 6=일
        days_until_tuesday = (1 - weekday) % 7
        this_tuesday = today + timedelta(days=days_until_tuesday)

        month = this_tuesday.month

        # 이번 달에서 몇 번째 화요일인지 계산
        first_day = this_tuesday.replace(day=1)
        first_tuesday_offset = (1 - first_day.weekday()) % 7
        first_tuesday = first_day + timedelta(days=first_tuesday_offset)
        week_number = ((this_tuesday - first_tuesday).days // 7) + 1

        return month, week_number

    def _get_output_filename(self) -> str:
        """출력 파일명 생성"""
        month, week_number = self._get_week_info()
        cafe_count = len(self.account_info.assigned_cafes)
        suffix = "_9개카페" if cafe_count > 1 else ""
        return f"(부시기획) LG전자 베스트샵 카페 모니터링_{month}월 {week_number}주차{suffix}.xlsx"

    def _normalize_text(self, text: str) -> str:
        """텍스트 정규화"""
        if not text:
            return ""
        text = text.replace('\n', ' ').replace('\r', ' ')
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def _wait(self, milliseconds: int):
        """밀리초 단위 대기"""
        time.sleep(milliseconds / 1000.0)

    def _wait_for_element(self, frame: FrameLike, selector: str, timeout: int = None) -> bool:
        """요소가 나타날 때까지 대기"""
        try:
            timeout = timeout or self.timeouts.ELEMENT_WAIT
            frame.wait_for_selector(selector, timeout=timeout, state='visible')
            return True
        except PlaywrightTimeoutError:
            return False

    def _find_element_with_selectors(
        self,
        frame: FrameLike,
        selectors: tuple,
        description: str = "요소",
        wait: bool = True
    ) -> Optional[Any]:
        """여러 선택자로 요소 찾기"""
        for selector in selectors:
            if wait:
                self._wait_for_element(frame, selector, timeout=2000)

            elem = frame.locator(selector).first
            if elem.count() > 0:
                self.logger.debug(f"{description} 발견: {selector}")
                return elem
        return None

    def _extract_article_id(self, url: str) -> Optional[str]:
        """URL에서 게시글 ID 추출"""
        match = re.search(r'/articles/(\d+)', url)
        return match.group(1) if match else None

    def _find_iframe(self, patterns: List[str], frame_name: str = None) -> Optional[Frame]:
        """iframe 찾기"""
        frames = self.page.frames

        for frame in frames:
            # name 우선 확인
            if frame_name:
                try:
                    if frame.name == frame_name:
                        self.logger.debug(f"iframe 발견 (name): {frame_name}")
                        return frame
                except:
                    pass

            # URL 패턴 확인
            frame_url = frame.url
            for pattern in patterns:
                if pattern in frame_url:
                    self.logger.debug(f"iframe 발견 (패턴: {pattern})")
                    return frame

        return None

    def _find_matching_inner_iframe(self, parent_frame: Frame, article_id: str) -> Optional[Frame]:
        """중첩 iframe 중 현재 게시글 ID와 일치하는 것 찾기"""
        inner_frames = parent_frame.child_frames

        if len(inner_frames) == 0:
            return None

        self.logger.debug(f"내부 iframe {len(inner_frames)}개 발견, 게시글 ID: {article_id}")

        for inner_frame in inner_frames:
            frame_url = inner_frame.url

            if '/ca-fe/cafes/' in frame_url and '/articles/' in frame_url:
                iframe_article_id = self._extract_article_id(frame_url)

                if iframe_article_id == article_id:
                    self.logger.debug(f"일치하는 iframe 발견: {frame_url}")
                    return inner_frame

        return None

    @retry(
        stop=stop_after_attempt(RetryConfig.MAX_ATTEMPTS),
        wait=wait_exponential(min=RetryConfig.MIN_WAIT, max=RetryConfig.MAX_WAIT),
        retry=retry_if_exception_type((PlaywrightTimeoutError, Exception)),
        reraise=True
    )
    def _collect_comments(self, article_frame: FrameLike, url: str) -> List[str]:
        """댓글 수집 (재시도 포함)"""
        comments = []

        # 댓글 버튼 클릭
        try:
            self.logger.debug("댓글 버튼 클릭 시도")
            comment_button = self._find_element_with_selectors(
                article_frame, self.selectors.COMMENT_BUTTON, "댓글 버튼", wait=True
            )

            if comment_button:
                comment_button.click()
                self._wait(self.wait_times.AFTER_COMMENT_CLICK)
                self.logger.debug("댓글 버튼 클릭 완료")
            else:
                self.logger.debug("댓글 버튼 없음, 스크롤 시도")
                for _ in range(self.constants.MAX_SCROLL_ATTEMPTS):
                    article_frame.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                    self._wait(self.wait_times.SCROLL_INTERVAL)
        except Exception as e:
            self.logger.debug(f"댓글 버튼 클릭 실패: {e}, 스크롤로 대체")
            for _ in range(self.constants.MAX_SCROLL_ATTEMPTS):
                article_frame.evaluate('window.scrollTo(0, document.body.scrollHeight)')
                self._wait(self.wait_times.SCROLL_INTERVAL)

        # 댓글 수집
        try:
            # 댓글 로딩 대기
            if not self._wait_for_element(article_frame, self.selectors.COMMENT_ITEMS[0], timeout=self.timeouts.COMMENT_LOAD):
                self.logger.debug(f"댓글 없음: {url}")
                return comments

            comment_count = article_frame.locator(self.selectors.COMMENT_ITEMS[0]).count()
            self.logger.debug(f"댓글 {comment_count}개 수집 시작")

            for i in range(comment_count):
                try:
                    comment_elem = article_frame.locator(self.selectors.COMMENT_ITEMS[0]).nth(i)

                    author_elem = comment_elem.locator(self.selectors.COMMENT_AUTHOR).first
                    author = author_elem.inner_text().strip() if author_elem.count() > 0 else "익명"

                    text_elem = comment_elem.locator(self.selectors.COMMENT_TEXT).first
                    text = self._normalize_text(text_elem.inner_text()) if text_elem.count() > 0 else ""

                    if text:
                        comments.append(f"{author} : {text}")

                except Exception as e:
                    self.logger.debug(f"댓글 {i} 수집 실패: {e}")
                    continue

            self.logger.debug(f"댓글 {len(comments)}개 수집 완료")

        except Exception as e:
            self.logger.debug(f"댓글 수집 실패: {e}")

        return comments

    @contextmanager
    def browser_context(self):
        """브라우저 컨텍스트 매니저"""
        try:
            self._start_browser()
            yield self.page
        finally:
            self._close_browser()

    def _start_browser(self):
        """브라우저 시작"""
        self.logger.info("브라우저 시작")

        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=False,
            args=[
                '--start-maximized',
                '--disable-blink-features=AutomationControlled',
                '--disable-dev-shm-usage',
                '--disable-gpu',  # GPU 메모리 사용 최소화
                '--disable-software-rasterizer',  # 소프트웨어 렌더링 비활성화
                '--disable-extensions',  # 확장 프로그램 비활성화
                '--no-sandbox',  # 샌드박스 비활성화 (메모리 절약)
                '--disable-setuid-sandbox',
                '--disable-features=TranslateUI',  # 번역 기능 비활성화
                '--disable-features=Translate',
                '--js-flags=--expose-gc'  # JavaScript 가비지 컬렉션 활성화
            ]
        )

        self.context = self.browser.new_context(
            viewport=None,
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        self.page = self.context.new_page()
        self.logger.info("브라우저 전체화면으로 시작됨")

    def _close_browser(self):
        """브라우저 종료 (메모리 정리 포함)"""
        try:
            # 페이지 리소스 정리
            if self.page:
                try:
                    self.page.evaluate('() => { window.stop(); }')
                    self.page.evaluate('() => { console.clear(); }')
                except:
                    pass

            # 컨텍스트 정리
            if self.context:
                try:
                    self.context.close()
                except:
                    pass

            # 브라우저 종료
            if self.browser:
                self.browser.close()
                self.logger.info("브라우저 종료")

            # Playwright 정리
            if self.playwright:
                self.playwright.stop()

            # Python 가비지 컬렉션 강제 실행
            import gc
            gc.collect()

        except Exception as e:
            self.logger.warning(f"브라우저 종료 중 오류: {e}")

    def _get_cookie_path(self) -> Path:
        """쿠키 파일 경로 반환"""
        return Path(self.config.log_folder) / f"cookies_{self.group_name}.json"

    def _save_cookies(self):
        """브라우저 쿠키 저장"""
        try:
            cookies = self.context.cookies()
            cookie_path = self._get_cookie_path()
            with open(cookie_path, 'w', encoding='utf-8') as f:
                json.dump(cookies, f)
            self.logger.info(f"쿠키 저장 완료: {cookie_path}")
        except Exception as e:
            self.logger.warning(f"쿠키 저장 실패: {e}")

    def _load_cookies(self) -> bool:
        """저장된 쿠키 로드"""
        try:
            cookie_path = self._get_cookie_path()
            if not cookie_path.exists():
                self.logger.info("저장된 쿠키 없음")
                return False

            with open(cookie_path, 'r', encoding='utf-8') as f:
                cookies = json.load(f)

            # 쿠키 만료 여부를 파일에서 직접 확인 (브라우저 접속 없음)
            now = time.time()
            auth_cookies = [c for c in cookies if c.get('name') in ('NID_AUT', 'NID_SES')]
            if not auth_cookies:
                self.logger.warning("인증 쿠키 없음 (NID_AUT/NID_SES)")
                return False
            for c in auth_cookies:
                expires = c.get('expires', 0)
                if expires and expires < now:
                    self.logger.warning(f"쿠키 만료됨: {c['name']} (만료: {datetime.fromtimestamp(expires)})")
                    return False

            self.context.add_cookies(cookies)
            self.logger.info("쿠키 로드 완료 (유효)")

        except Exception as e:
            self.logger.warning(f"쿠키 로드 실패: {e}")
            return False

    def _load_existing_urls(self):
        """엑셀 파일에서 기존 URL 로드"""
        try:
            output_folder = Path(self.config.output_folder)
            filename = self._get_output_filename()
            filepath = output_folder / filename

            if not filepath.exists():
                self.logger.info("기존 엑셀 파일 없음, 처음부터 시작")
                self.existing_urls = set()
                return

            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active

            # URL 컬럼 찾기 (헤더 행에서 'URL' 찾기)
            url_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'URL':
                    url_col = col_idx
                    break

            if url_col:
                # 모든 URL 수집 (헤더 제외)
                urls = set()
                for row in ws.iter_rows(min_row=2, min_col=url_col, max_col=url_col):
                    url = row[0].value
                    if url:
                        urls.add(url)

                self.existing_urls = urls
                self.logger.info(f"기존 URL {len(urls)}개 로드 완료")
            else:
                self.logger.warning("엑셀 파일에 URL 컬럼 없음")
                self.existing_urls = set()

            wb.close()

        except Exception as e:
            self.logger.warning(f"기존 URL 로드 실패: {e}")
            self.existing_urls = set()

    def _should_restart_browser(self) -> bool:
        """브라우저 재시작 필요 여부 확인"""
        elapsed = time.time() - self.last_restart_time
        return elapsed >= self.restart_interval

    def _restart_browser_if_needed(self):
        """필요시 브라우저 재시작"""
        if not self._should_restart_browser():
            return False

        elapsed_minutes = int((time.time() - self.last_restart_time) / 60)
        self.restart_count += 1

        self.logger.info("="*80)
        self.logger.info(f"브라우저 재시작 #{self.restart_count} (경과: {elapsed_minutes}분)")
        self.logger.info("="*80)
        print(f"\n{'='*80}")
        print(f"🔄 [{self.group_name}] 브라우저 재시작 #{self.restart_count} (메모리 최적화)")
        print(f"   경과 시간: {elapsed_minutes}분")
        print(f"{'='*80}\n")

        # 브라우저 종료
        self._close_browser()
        self._wait(2000)  # 2초 대기

        # 브라우저 재시작
        self._start_browser()

        # 쿠키 로드 시도
        cookie_loaded = self._load_cookies()

        if not cookie_loaded:
            # 쿠키 로드 실패 시 재로그인
            self.logger.info("쿠키 로드 실패, 재로그인 필요")
            print(f"[{self.group_name}] 재로그인 필요...\n")
            if not self.login_naver():
                raise Exception("재로그인 실패")
            self._save_cookies()
        else:
            self.logger.info("쿠키로 로그인 생략")
            print(f"[{self.group_name}] 쿠키 재사용 (로그인 생략)\n")

        # 기존 URL 다시 로드
        self._load_existing_urls()

        # 타이머 리셋
        self.last_restart_time = time.time()

        return True

    def login_naver(self) -> bool:
        """네이버 로그인"""
        self.logger.info("네이버 로그인 시작")

        try:
            self.page.goto('https://nid.naver.com/nidlogin.login', wait_until='domcontentloaded')
            self._wait(self.wait_times.AFTER_PAGE_LOAD)

            # 아이디/비밀번호 입력 (사람처럼 한 글자씩 타이핑)
            self.logger.debug(f"아이디 입력: {self.account_info.naver_id}")
            self.page.click(self.selectors.LOGIN_ID)
            self._wait(random.randint(300, 600))
            self.page.type(self.selectors.LOGIN_ID, self.account_info.naver_id, delay=random.randint(80, 160))
            self._wait(random.randint(400, 800))

            self.logger.debug("비밀번호 입력")
            self.page.click(self.selectors.LOGIN_PW)
            self._wait(random.randint(300, 600))
            self.page.type(self.selectors.LOGIN_PW, self.account_info.naver_password, delay=random.randint(80, 160))
            self._wait(random.randint(400, 800))

            # 로그인 버튼 클릭
            self.logger.debug("로그인 버튼 클릭")
            try:
                login_button = self.page.locator(self.selectors.LOGIN_BUTTON)
                if login_button.count() > 0:
                    login_button.click()
                else:
                    self.page.click(self.selectors.LOGIN_BUTTON_ALT)
            except Exception as e:
                self.logger.debug(f"버튼 클릭 실패, Enter 키 사용: {e}")
                self.page.press(self.selectors.LOGIN_PW, 'Enter')

            self._wait(self.wait_times.AFTER_LOGIN)

            # 로그인 오류 확인
            try:
                error_msg = self.page.locator(self.selectors.LOGIN_ERROR).first
                if error_msg.count() > 0:
                    error_text = error_msg.inner_text()
                    self.logger.error(f"로그인 오류: {error_text}")
                    print(f"\n❌ 로그인 오류: {error_text}\n")

                screenshot_path = Path(self.config.log_folder) / f'login_{datetime.now().strftime("%Y%m%d_%H%M%S")}.png'
                self.page.screenshot(path=str(screenshot_path))
            except Exception as e:
                self.logger.debug(f"스크린샷 실패: {e}")

            # 로그인 완료 대기
            self.logger.info("로그인 완료 대기 (최대 120초)")
            print("\n" + "="*80)
            print("⚠️  로그인 처리 중...")
            print("    - 보안인증이 필요하면 브라우저에서 직접 처리해주세요")
            print("="*80 + "\n")

            max_checks = self.constants.LOGIN_TIMEOUT_SECONDS // self.constants.LOGIN_CHECK_INTERVAL
            for i in range(max_checks):
                self._wait(self.constants.LOGIN_CHECK_INTERVAL * 1000)
                current_url = self.page.url

                if 'nid.naver.com/nidlogin' not in current_url:
                    elapsed = (i+1) * self.constants.LOGIN_CHECK_INTERVAL
                    self.logger.info(f"로그인 성공 (소요: {elapsed}초)")
                    print(f"\n✅ 로그인 성공 (소요: {elapsed}초)\n")
                    
                    # 쿠키 저장
                    self._save_cookies()
                    
                    return True

                remaining = self.constants.LOGIN_TIMEOUT_SECONDS - (i+1) * self.constants.LOGIN_CHECK_INTERVAL
                print(f"[{(i+1) * self.constants.LOGIN_CHECK_INTERVAL}초] 대기 중... (남은 시간: {remaining}초)")

            # 타임아웃
            if 'nid.naver.com/nidlogin' in self.page.url:
                self.logger.error("로그인 실패: 타임아웃")
                print("\n❌ 로그인 실패: 타임아웃\n")
                return False

            return True

        except Exception as e:
            self.logger.error(f"로그인 오류: {e}")
            return False

    @retry(
        stop=stop_after_attempt(RetryConfig.MAX_ATTEMPTS),
        wait=wait_exponential(min=RetryConfig.MIN_WAIT, max=RetryConfig.MAX_WAIT),
        retry=retry_if_exception_type((PlaywrightTimeoutError,)),
        reraise=True
    )
    def search_keyword_in_cafe(self, cafe_id: str, keyword: str, page_num: int) -> List[Dict[str, Any]]:
        """카페 내 키워드 검색 (재시도 포함)"""
        self.logger.debug(f"'{keyword}' {page_num}페이지 검색")

        posts = []

        try:
            encoded_keyword = quote(keyword)
            search_url = f"https://cafe.naver.com/f-e/cafes/{cafe_id}/menus/0?viewType=L&ta=ARTICLE_COMMENT&page={page_num}&q={encoded_keyword}&p=7d"

            self.page.goto(search_url, wait_until='domcontentloaded', timeout=self.timeouts.PAGE_LOAD)
            self._wait(self.wait_times.AFTER_PAGE_LOAD)

            # 검색 결과 iframe 찾기
            search_frame = self._find_iframe(['ArticleSearchList', 'menus'])

            if not search_frame:
                self.logger.debug("검색 iframe 미발견, 메인 페이지 사용")
                search_frame = self.page

            # 게시글 링크 찾기
            link_elements = self._find_element_with_selectors(
                search_frame, self.selectors.ARTICLE_LINKS, "게시글 링크"
            )

            if not link_elements:
                self.logger.debug(f"'{keyword}' {page_num}페이지 결과 없음")
                return posts

            link_count = search_frame.locator(self.selectors.ARTICLE_LINKS[0]).count()
            self.logger.debug(f"{link_count}개 게시글 발견")

            # URL 수집
            for i in range(link_count):
                try:
                    link = search_frame.locator(self.selectors.ARTICLE_LINKS[0]).nth(i)
                    href = link.get_attribute('href')

                    if href and '/articles/' in href:
                        if href.startswith('http'):
                            post_url = href
                        elif href.startswith('/'):
                            post_url = f"https://cafe.naver.com{href}"
                        else:
                            post_url = f"https://cafe.naver.com/{href}"

                        posts.append({'keyword': keyword, 'url': post_url})

                except Exception as e:
                    self.logger.debug(f"URL 수집 오류 (인덱스 {i}): {e}")
                    continue

        except Exception as e:
            self.logger.warning(f"'{keyword}' {page_num}페이지 검색 오류: {e}")
            raise

        # 메모리 정리
        try:
            self.page.evaluate('() => { if (window.gc) window.gc(); }')
        except:
            pass

        self.logger.debug(f"'{keyword}' {page_num}페이지: {len(posts)}개 URL 수집")
        return posts

    @retry(
        stop=stop_after_attempt(RetryConfig.MAX_ATTEMPTS),
        wait=wait_exponential(min=RetryConfig.MIN_WAIT, max=RetryConfig.MAX_WAIT),
        retry=retry_if_exception_type((PlaywrightTimeoutError,)),
        reraise=True
    )
    def collect_post_details(self, post_info: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """게시글 상세 정보 수집 (재시도 포함)"""
        url = post_info['url']
        keyword = post_info['keyword']
        cafe_name = post_info['cafe_name']

        # 중복 URL 체크
        if url in self.existing_urls:
            self.logger.debug(f"중복 URL 건너뛰기: {url}")
            return None

        try:
            # 게시글 페이지 이동
            self.page.goto(url, wait_until='domcontentloaded', timeout=self.timeouts.PAGE_LOAD)
            self._wait(self.wait_times.AFTER_PAGE_LOAD)

            # 추가 대기 - 동적 콘텐츠 로딩을 위해
            self._wait(2000)

            # 모든 프레임에서 요소 찾기 시도
            article_frame = None
            for frame in self.page.frames:
                try:
                    # 제목이 있는 프레임을 찾음
                    if frame.locator('h3.title_text').count() > 0:
                        article_frame = frame
                        self.logger.debug(f"게시글 프레임 발견: {frame.url}")
                        break
                except:
                    continue

            if not article_frame:
                self.logger.warning(f"게시글 프레임 미발견, 메인 페이지 사용 ({url})")
                self.logger.debug(f"사용 가능한 프레임 수: {len(self.page.frames)}")
                for idx, frame in enumerate(self.page.frames):
                    self.logger.debug(f"프레임 {idx}: {frame.url}")
                article_frame = self.page

            # 기본 정보 수집
            try:
                # 제목
                title_elem = article_frame.locator(self.selectors.TITLE).first
                title = self._normalize_text(title_elem.inner_text()) if title_elem.count() > 0 else ""
                self.logger.debug(f"제목: {title[:50] if title else 'None'}...")

                # 작성자
                author_elem = article_frame.locator(self.selectors.AUTHOR).first
                author = author_elem.inner_text().strip() if author_elem.count() > 0 else ""
                self.logger.debug(f"작성자: {author if author else 'None'}")

                # 날짜
                date_elem = article_frame.locator(self.selectors.DATE).first
                date_str = date_elem.inner_text().strip() if date_elem.count() > 0 else ""
                self.logger.debug(f"날짜: {date_str if date_str else 'None'}")

                # 본문 내용
                content_elem = article_frame.locator(self.selectors.CONTENT).first
                content = self._normalize_text(content_elem.inner_text()) if content_elem.count() > 0 else ""
                self.logger.debug(f"내용 길이: {len(content)} 글자")

                # 좋아요 수 추출 (여러 선택자 시도)
                like_elem = self._find_element_with_selectors(article_frame, self.selectors.LIKES, "좋아요", wait=False)
                likes = "0"
                if like_elem and like_elem.count() > 0:
                    likes_text = like_elem.inner_text().strip()
                    likes = re.sub(r'\D', '', likes_text) if likes_text else "0"
                    if not likes:  # 숫자가 없으면 0으로
                        likes = "0"
                self.logger.debug(f"좋아요: {likes}")

            except Exception as e:
                self.logger.error(f"기본 정보 수집 실패 ({url}): {e}")
                self.logger.error(f"상세 오류: {traceback.format_exc()}")
                return None

            # 댓글 수집 (중첩 iframe 처리 제거 - Frame detached 오류 방지)
            comments = []
            try:
                comments = self._collect_comments(article_frame, url)
            except Exception as comment_err:
                self.logger.debug(f"댓글 수집 실패: {comment_err}")

            if len(comments) == 0:
                self.logger.debug(f"댓글 없음: {url}")
            else:
                self.logger.debug(f"댓글 {len(comments)}개 수집 완료")

            # 수집 완료 후 existing_urls에 추가
            self.existing_urls.add(url)

            # 메모리 정리 (더 적극적)
            try:
                # JavaScript 가비지 컬렉션 실행
                self.page.evaluate('() => { if (window.gc) window.gc(); }')
                # 페이지 리소스 정리
                self.page.evaluate('() => { window.stop(); }')
                # 콘솔 로그 정리
                self.page.evaluate('() => { console.clear(); }')
            except:
                pass

            return {
                '채널': cafe_name,
                '키워드': keyword,
                '닉네임': author,
                '날짜': date_str,
                '제목': title,
                '내용': content,
                '좋아요': likes,
                'URL': url,
                '댓글': comments
            }

        except Exception as e:
            self.logger.warning(f"게시글 수집 오류 ({url}): {e}")
            raise

    def _save_batch_to_excel(self, keyword: str):
        """키워드별 배치 저장"""
        if len(self.collected_data) == 0:
            self.logger.info(f"'{keyword}' 수집 데이터 없음, 저장 건너뜀")
            return

        self.logger.info(f"'{keyword}' {len(self.collected_data)}개 데이터 저장 중...")

        # 출력 폴더 생성
        output_folder = Path(self.config.output_folder)
        output_folder.mkdir(parents=True, exist_ok=True)

        # 파일명 생성
        filename = self._get_output_filename()
        filepath = output_folder / filename

        # 기존 파일이 있으면 로드, 없으면 새로 생성
        if filepath.exists():
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            existing_rows = ws.max_row
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "모니터링"
            existing_rows = 0

            # 헤더 생성 (새 파일인 경우)
            max_comments = max((len(data.get('댓글', [])) for data in self.collected_data), default=0)
            headers = ['채널', '키워드', '닉네임', '날짜', '제목', '내용', '좋아요', 'URL']
            headers.extend([f'댓글{i}' for i in range(1, max_comments + 1)])

            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 현재 최대 댓글 컬럼 수 확인
        current_max_cols = ws.max_column
        new_max_comments = max((len(data.get('댓글', [])) for data in self.collected_data), default=0)
        required_cols = 8 + new_max_comments  # 기본 8개 + 댓글 컬럼

        # 필요시 댓글 컬럼 추가
        if required_cols > current_max_cols:
            for i in range(current_max_cols - 7, new_max_comments + 1):
                ws.cell(row=1, column=8 + i, value=f'댓글{i}')
                ws.cell(row=1, column=8 + i).font = Font(bold=True)
                ws.cell(row=1, column=8 + i).alignment = Alignment(horizontal='center', vertical='center')

        # 데이터 추가
        for data in self.collected_data:
            row = [
                data.get('채널', ''),
                data.get('키워드', ''),
                data.get('닉네임', ''),
                data.get('날짜', ''),
                data.get('제목', ''),
                data.get('내용', ''),
                data.get('좋아요', '0'),
                data.get('URL', '')
            ]

            comments = data.get('댓글', [])
            row.extend(comments)

            # 남은 컬럼 빈칸
            max_cols = ws.max_column
            row.extend([''] * (max_cols - len(row)))

            ws.append(row)

        # 컬럼 너비 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 저장
        wb.save(filepath)
        self.logger.info(f"'{keyword}' 저장 완료: {filepath}")
        print(f"  💾 '{keyword}' {len(self.collected_data)}개 저장: {filepath}")

        # 메모리 해제
        self.collected_data.clear()

    def _crawl_keyword(self, cafe_id: str, cafe_name: str, keyword: str, keyword_idx: int, total_keywords: int):
        """단일 키워드 크롤링"""
        print(f"\n{'='*80}")
        print(f"[{cafe_name}] [키워드 {keyword_idx}/{total_keywords}] '{keyword}' 검색 시작")
        print(f"{'='*80}")
        self.logger.info(f"[{cafe_name}] [{keyword_idx}/{total_keywords}] '{keyword}' 시작")

        page_num = 1
        keyword_total_posts = 0

        while True:
            print(f"\n[{cafe_name}] [{keyword}] {page_num}페이지 검색 중...")

            try:
                # 게시글 URL 수집
                posts = self.search_keyword_in_cafe(cafe_id, keyword, page_num)

                # 각 post에 cafe_name 추가
                for post in posts:
                    post['cafe_name'] = cafe_name

                if len(posts) == 0:
                    print(f"  → {page_num}페이지 게시글 없음. '{keyword}' 종료")
                    self.logger.info(f"'{keyword}' {page_num}페이지 없음, 종료")
                    break

                print(f"  → {len(posts)}개 발견")
                self.logger.info(f"'{keyword}' {page_num}페이지: {len(posts)}개")

                # 각 게시글 처리
                page_collected = 0
                for post_idx, post_info in enumerate(posts, 1):
                    print(f"  [{post_idx}/{len(posts)}] 처리 중...")

                    # 게시글 처리 전 브라우저 재시작 체크 (더 자주 체크)
                    try:
                        restarted = self._restart_browser_if_needed()
                        if restarted:
                            self.logger.info(f"브라우저 재시작 후 게시글 처리 계속")
                    except Exception as e:
                        self.logger.error(f"브라우저 재시작 실패: {e}")
                        raise

                    try:
                        post_data = self.collect_post_details(post_info)

                        if post_data:
                            self.collected_data.append(post_data)
                            page_collected += 1
                            keyword_total_posts += 1
                            comment_count = len(post_data.get('댓글', []))
                            print(f"    ✅ 완료 (댓글 {comment_count}개)")
                        else:
                            print(f"    ⏭️  수집 실패, 건너뜀")

                    except Exception as e:
                        self.logger.warning(f"게시글 처리 실패 ({post_info['url']}): {e}")
                        print(f"    ❌ 오류, 건너뜀")
                        continue

                print(f"\n  {page_num}페이지 완료: {page_collected}개")

                # 다음 페이지로
                page_num += 1

                # 브라우저 재시작 체크
                try:
                    restarted = self._restart_browser_if_needed()
                    if restarted:
                        self.logger.info(f"브라우저 재시작 후 '{keyword}' 계속")
                except Exception as e:
                    self.logger.error(f"브라우저 재시작 실패: {e}")
                    raise

                # Rate limiting: 랜덤 대기 시간 추가
                random_wait = random.uniform(500, 2000)
                self._wait(self.wait_times.BETWEEN_PAGES + random_wait)

            except Exception as e:
                self.logger.error(f"'{keyword}' {page_num}페이지 오류: {e}")
                print(f"  ❌ 오류 발생, 다음 키워드로")
                break

        print(f"\n'{keyword}' 완료: 총 {keyword_total_posts}개")
        self.logger.info(f"'{keyword}' 완료: {keyword_total_posts}개")

        # 키워드별 배치 저장
        self._save_batch_to_excel(keyword)

    def _setup(self):
        """초기 설정 (로그인만)"""
        self.logger.info("="*80)
        self.logger.info("네이버 카페 크롤링 시작")
        self.logger.info("="*80)

        # 기존 URL 로드
        self._load_existing_urls()

        # 로그인
        if not self.login_naver():
            raise Exception("로그인 실패")

    def _crawl_cafe(self, cafe: CafeInfo, cafe_idx: int, total_cafes: int):
        """단일 카페 크롤링"""
        print(f"\n{'='*80}")
        print(f"[카페 {cafe_idx}/{total_cafes}] {cafe.cafe_name} 크롤링 시작")
        print(f"{'='*80}")
        self.logger.info(f"[카페 {cafe_idx}/{total_cafes}] {cafe.cafe_name} 시작")

        # 모든 키워드 크롤링
        keywords = self.config.keywords
        total_keywords = len(keywords)

        for keyword_idx, keyword in enumerate(keywords, 1):
            try:
                self._crawl_keyword(cafe.cafe_id, cafe.cafe_name, keyword, keyword_idx, total_keywords)
            except Exception as e:
                self.logger.error(f"[{cafe.cafe_name}] '{keyword}' 크롤링 실패: {e}")
                print(f"\n❌ [{cafe.cafe_name}] '{keyword}' 크롤링 실패, 다음 키워드로 이동\n")
                continue

        print(f"\n{'='*80}")
        print(f"[카페 {cafe_idx}/{total_cafes}] {cafe.cafe_name} 완료")
        print(f"{'='*80}")
        self.logger.info(f"[카페 {cafe_idx}/{total_cafes}] {cafe.cafe_name} 완료")

    def run(self):
        """메인 실행"""
        try:
            with self.browser_context():
                self._setup()

                # 담당 카페만 필터링
                assigned_cafe_names = self.account_info.assigned_cafes
                cafes = [cafe for cafe in self.config.cafes if cafe.cafe_name in assigned_cafe_names]
                total_cafes = len(cafes)

                for cafe_idx, cafe in enumerate(cafes, 1):
                    try:
                        self._crawl_cafe(cafe, cafe_idx, total_cafes)
                    except Exception as e:
                        self.logger.error(f"카페 크롤링 실패 ({cafe.cafe_name}): {e}")
                        print(f"\n❌ 카페 크롤링 실패 ({cafe.cafe_name}), 다음 카페로 이동\n")
                        continue

            self.logger.info("="*80)
            self.logger.info("크롤링 완료!")
            self.logger.info("="*80)
            print(f"\n{'='*80}")
            print("✅ 전체 크롤링 완료!")
            print(f"{'='*80}\n")

        except Exception as e:
            self.logger.error(f"치명적 오류: {e}")
            self.logger.error(traceback.format_exc())
            print(f"\n❌ 치명적 오류: {e}\n")


def run_crawler_for_account(account_info_dict: dict, config_path: str):
    """각 계정별로 크롤러를 실행하는 프로세스 함수"""
    try:
        account_info = AccountConfig(**account_info_dict)
        group_name = account_info.group_name

        print(f"\n[{group_name}] 프로세스 시작 (계정: {account_info.naver_id})")
        print(f"[{group_name}] 담당 카페: {', '.join(account_info.assigned_cafes)}\n")

        crawler = NaverCafeCrawler(
            config_path=config_path,
            account_info=account_info,
            group_name=group_name
        )
        crawler.run()

        print(f"\n[{group_name}] 프로세스 완료\n")
    except Exception as e:
        print(f"\n❌ [{group_name}] 프로세스 오류: {e}\n")
        traceback.print_exc()


def main():
    """메인 함수 - 멀티프로세싱 실행"""
    print("="*80)
    print("(부시기획) LG전자 베스트샵 카페 모니터링 크롤러")
    print("="*80)
    print()

    try:
        # 설정 로드
        with open("config.json", 'r', encoding='utf-8') as f:
            config_dict = json.load(f)

        config = CrawlerSettings(**config_dict)

        # 계정별로 프로세스 생성
        processes = []
        for account in config.accounts:
            account_dict = account.model_dump()
            p = multiprocessing.Process(
                target=run_crawler_for_account,
                args=(account_dict, "config.json")
            )
            processes.append(p)
            p.start()
            print(f"✅ [{account.group_name}] 프로세스 시작됨 (PID: {p.pid})")

        print(f"\n총 {len(processes)}개 프로세스 실행 중...\n")

        # 모든 프로세스 완료 대기
        for p in processes:
            p.join()

        print("\n" + "="*80)
        print("✅ 모든 크롤링 완료!")
        print("="*80)

    except Exception as e:
        print(f"\n❌ 초기화 실패: {e}\n")
        traceback.print_exc()
        return

    print("\n프로그램 종료")


if __name__ == "__main__":
    # Windows에서 multiprocessing 사용 시 필요
    multiprocessing.freeze_support()
    main()
