#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Instagram 댓글 크롤러
- 게시물 댓글 수집 (닉네임, 내용, 팔로우 여부)
- 팔로우 여부: 게시물 작성자의 팔로워 목록에서 검색
"""

import json
import sys
import time
import random
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Set

from playwright.sync_api import sync_playwright, Page, Browser
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Windows 콘솔 인코딩 설정
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8')


class InstagramCrawler:
    """Instagram 댓글 크롤러"""

    def __init__(self, config_path: str = "config.json"):
        """초기화"""
        self.config_path = Path(config_path)
        self.config = self._load_config()

        # 브라우저 관련
        self.playwright = None
        self.browser: Browser = None
        self.context = None
        self.page: Page = None

        # 쿠키 파일 경로
        self.cookie_path = self.config_path.parent / "cookies.json"

        # 수집된 댓글 데이터
        self.comments_data: List[Dict] = []

        # 팔로워 캐시 (이미 확인한 사용자)
        self.follower_cache: Dict[str, bool] = {}

    def _load_config(self) -> dict:
        """설정 파일 로드"""
        if not self.config_path.exists():
            raise FileNotFoundError(f"설정 파일을 찾을 수 없습니다: {self.config_path}")

        with open(self.config_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def _wait(self, seconds: float):
        """대기"""
        time.sleep(seconds)

    def _start_browser(self):
        """브라우저 시작"""
        print(f"\n{'='*80}")
        print("브라우저 시작")
        print(f"{'='*80}")

        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=False,
            args=[
                '--start-maximized',
                '--disable-blink-features=AutomationControlled',
                '--disable-dev-shm-usage',
            ]
        )

        self.context = self.browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        self.page = self.context.new_page()
        print(f"  → 브라우저 시작 완료")

    def _close_browser(self):
        """브라우저 종료"""
        try:
            if self.context:
                self.context.close()
            if self.browser:
                self.browser.close()
            if self.playwright:
                self.playwright.stop()
            print(f"\n{'='*80}")
            print("브라우저 종료 완료")
            print(f"{'='*80}")
        except Exception as e:
            print(f"  ❌ 브라우저 종료 중 오류: {e}")

    def _save_cookies(self):
        """쿠키 저장"""
        try:
            cookies = self.context.cookies()
            with open(self.cookie_path, 'w', encoding='utf-8') as f:
                json.dump(cookies, f)
            print(f"  💾 쿠키 저장 완료: {self.cookie_path}")
        except Exception as e:
            print(f"  ❌ 쿠키 저장 실패: {e}")

    def _load_cookies(self) -> bool:
        """쿠키 로드"""
        try:
            if not self.cookie_path.exists():
                print(f"  → 저장된 쿠키 없음")
                return False

            with open(self.cookie_path, 'r', encoding='utf-8') as f:
                cookies = json.load(f)

            self.context.add_cookies(cookies)
            print(f"  → 쿠키 로드 완료")

            # 쿠키 유효성 확인
            self.page.goto('https://www.instagram.com/', wait_until='domcontentloaded')
            self._wait(3)

            # 로그인 상태 확인
            try:
                profile_icon = self.page.locator('svg[aria-label="홈"]').first
                if profile_icon.count() > 0:
                    print(f"  ✅ 쿠키로 로그인 성공")
                    return True
            except:
                pass

            print(f"  → 쿠키 만료됨")
            return False

        except Exception as e:
            print(f"  ❌ 쿠키 로드 실패: {e}")
            return False

    def login(self) -> bool:
        """Instagram 로그인"""
        print(f"\n{'='*80}")
        print("Instagram 로그인")
        print(f"{'='*80}")

        # 쿠키로 로그인 시도
        if self._load_cookies():
            # 게시물 URL로 이동
            print(f"  → 게시물로 이동: {self.config['post_url']}")
            self.page.goto(self.config['post_url'], wait_until='domcontentloaded')
            self._wait(3)
            return True

        # 수동 로그인
        print(f"\n[로그인] 로그인 페이지로 이동...")
        self.page.goto('https://www.instagram.com/accounts/login/', wait_until='domcontentloaded')
        self._wait(3)

        try:
            # 아이디 입력
            print(f"  [1/3] 아이디 입력 중...")
            username_input = self.page.locator('input[name="username"]')
            username_input.fill(self.config['instagram_id'])
            self._wait(0.5)
            print(f"    ✅ 완료")

            # 비밀번호 입력
            print(f"  [2/3] 비밀번호 입력 중...")
            password_input = self.page.locator('input[name="password"]')
            password_input.fill(self.config['instagram_password'])
            self._wait(0.5)
            print(f"    ✅ 완료")

            # 로그인 버튼 클릭
            print(f"  [3/3] 로그인 버튼 클릭...")
            login_button = self.page.locator('button[type="submit"]')
            login_button.click()
            print(f"    ✅ 완료")

            # 10초 대기
            print(f"\n  → 로그인 처리 대기 중 (10초)...")
            self._wait(10)

            # 바로 게시물 URL로 이동
            print(f"  → 게시물로 이동: {self.config['post_url']}")
            self.page.goto(self.config['post_url'], wait_until='domcontentloaded')
            self._wait(3)

            self._save_cookies()
            print(f"\n  ✅ 로그인 및 이동 완료")
            return True

        except Exception as e:
            print(f"  ❌ 로그인 오류: {e}")
            return False

    def _scroll_until_hidden_comments(self):
        """숨겨진 댓글 보기 버튼이 나올 때까지 댓글창 스크롤"""
        print(f"\n{'='*80}")
        print("댓글창 스크롤 (숨겨진 댓글 보기 버튼 찾는 중)")
        print(f"{'='*80}")

        scroll_count = 0
        max_scrolls = 500  # 최대 스크롤 횟수

        # 댓글 컨테이너 스크롤 JavaScript
        scroll_comment_js = """
        () => {
            // 댓글 컨테이너 찾기 (클래스 기반)
            const commentContainer = document.querySelector('div.x5yr21d.xw2csxc.x1odjw0f.x1n2onr6');

            if (commentContainer) {
                commentContainer.scrollTop = commentContainer.scrollHeight;
                return true;
            }

            return false;
        }
        """

        while scroll_count < max_scrolls:
            # 숨겨진 댓글 보기 버튼 확인
            try:
                hidden_btn = self.page.locator('svg[aria-label="숨겨진 댓글 보기"]').first
                if hidden_btn.count() > 0:
                    print(f"\n  ✅ 숨겨진 댓글 보기 버튼 발견!")
                    break
            except:
                pass

            # 댓글창 스크롤
            scrolled = self.page.evaluate(scroll_comment_js)
            scroll_count += 1

            if scrolled:
                print(f"  [댓글창 스크롤] #{scroll_count}")
            else:
                print(f"  [스크롤] #{scroll_count} (댓글 컨테이너 못찾음)")

            self._wait(1)

        print(f"\n  → 스크롤 완료 (총 {scroll_count}회)")

    def _click_hidden_comments(self) -> int:
        """숨겨진 댓글 보기 버튼 클릭 (한 번만)"""
        print(f"\n{'='*80}")
        print("숨겨진 댓글 보기")
        print(f"{'='*80}")

        try:
            hidden_comments = self.page.locator('svg[aria-label="숨겨진 댓글 보기"]').first
            if hidden_comments.count() > 0:
                hidden_comments.click()
                print(f"  ✅ 숨겨진 댓글 보기 클릭 완료")
                self._wait(2)
                return 1
            else:
                print(f"  → 숨겨진 댓글 보기 버튼 없음")
                return 0
        except Exception as e:
            print(f"  → 숨겨진 댓글 보기 버튼 없음")
            return 0

    def _click_reply_buttons(self) -> int:
        """답글 보기 버튼 클릭"""
        print(f"\n{'='*80}")
        print("답글 펼치기")
        print(f"{'='*80}")

        click_count = 0
        max_attempts = 200  # 최대 시도 횟수

        for attempt in range(max_attempts):
            try:
                # "답글 n개 모두 보기" 버튼 찾기
                reply_buttons = self.page.locator('span.x1lliihq').filter(has_text="답글").filter(has_text="모두 보기").all()

                if not reply_buttons:
                    print(f"  → 더 이상 답글 보기 버튼 없음")
                    break

                clicked_this_round = False
                for btn in reply_buttons:
                    try:
                        if btn.is_visible():
                            btn.click()
                            click_count += 1
                            print(f"  [답글 보기] 클릭 #{click_count}")
                            self._wait(1)
                            clicked_this_round = True
                    except:
                        continue

                if not clicked_this_round:
                    break

            except:
                break

        print(f"\n  → 답글 펼치기 완료 (총 {click_count}회 클릭)")
        return click_count

    def _extract_comments(self) -> List[Dict]:
        """댓글 추출"""
        print(f"\n{'='*80}")
        print("댓글 수집")
        print(f"{'='*80}")

        # JavaScript로 댓글 추출
        js_code = """
        () => {
            const results = [];
            const seen = new Set();

            document.querySelectorAll('span[dir="auto"]').forEach(span => {
                const text = span.textContent.trim();
                const style = span.getAttribute('style') || '';

                // 필터링 조건
                const isUsername = /^[a-zA-Z0-9_.]+$/.test(text);
                const uiTexts = ['Meta의 다른 앱', '좋아요', '답글 달기', '번역 보기', '탐색 탭', '더 보기', '메시지', '만들기', '프로필', '블로그', '채용 정보', '도움말', '개인정보처리방침', 'Meta AI', '한국어', 'Instagram Lite', 'Meta Verified'];
                const isUI = uiTexts.includes(text) || text.includes('연락처 업로드') || text.includes('© 2026') || text.includes('© 2025') || text.includes('Instagram from');
                const hasReply = (text.includes('답글') && text.includes('보기')) || (text.includes('답글') && text.includes('숨기기'));
                const isTime = /^\\d+[시분초주일개월년]/.test(text) || /^\\d+[시분초주일개월년]\\s*전?$/.test(text);
                const isLike = /^좋아요\\s*\\d*개?$/.test(text);

                // 댓글 내용 조건: 3자 이상, line-clamp 스타일, 필터링 통과
                if (text.length >= 3 && style.includes('line-clamp') && !isUsername && !isUI && !text.startsWith('@') && !hasReply && !isTime && !isLike) {

                    // 부모를 거슬러 올라가며 닉네임과 시간 찾기
                    let username = '';
                    let datetime = '';
                    let el = span;
                    for (let level = 0; level < 15; level++) {
                        if (!el.parentElement) break;
                        el = el.parentElement;

                        // 닉네임 찾기
                        if (!username) {
                            const links = el.querySelectorAll('a[href]');
                            for (const a of links) {
                                const href = a.getAttribute('href');
                                if (href && href.match(/^\\/[a-zA-Z0-9_.]+\\/$/) && !['reels', 'explore', ''].includes(href.replace(/\\//g, ''))) {
                                    username = href.replace(/\\//g, '');
                                    break;
                                }
                            }
                        }

                        // 시간 찾기
                        if (!datetime) {
                            const timeTag = el.querySelector('time[datetime]');
                            if (timeTag) {
                                datetime = timeTag.getAttribute('datetime');
                            }
                        }

                        if (username && datetime) break;
                    }

                    // 중복 체크
                    const key = username + ':' + text.substring(0, 30);
                    if (username && !seen.has(key)) {
                        seen.add(key);
                        results.push({
                            username: username,
                            content: text,
                            datetime: datetime,
                            is_reply: false
                        });
                    }
                }
            });

            return results;
        }
        """

        try:
            comments = self.page.evaluate(js_code)
            print(f"  → {len(comments)}개 댓글 추출 완료")
            return comments
        except Exception as e:
            print(f"  ❌ 댓글 추출 오류: {e}")
            return []

    def _check_follow_status(self, usernames: List[str]) -> Dict[str, bool]:
        """팔로우 여부 확인"""
        print(f"\n{'='*80}")
        print(f"팔로우 여부 확인 ({len(usernames)}명)")
        print(f"{'='*80}")

        results = {}
        post_author = self.config['post_author']
        total = len(usernames)

        try:
            # 작성자 프로필로 이동
            print(f"\n[이동] {post_author} 프로필...")
            self.page.goto(f'https://www.instagram.com/{post_author}/', wait_until='domcontentloaded')
            self._wait(2)
            print(f"  ✅ 완료")

            # 팔로워 버튼 클릭
            print(f"\n[클릭] 팔로워 목록 열기...")
            follower_link = self.page.locator(f'a[href="/{post_author}/followers/"]').first
            if follower_link.count() == 0:
                follower_link = self.page.locator('a:has-text("팔로워")').first

            if follower_link.count() == 0:
                print(f"  ❌ 팔로워 버튼을 찾을 수 없습니다")
                return {u: False for u in usernames}

            follower_link.click()
            self._wait(2)
            print(f"  ✅ 완료")

            # 검색창 찾기
            search_input = self.page.locator('input[placeholder="검색"]').first
            if search_input.count() == 0:
                search_input = self.page.locator('input[type="text"]').first

            if search_input.count() == 0:
                print(f"  ❌ 검색창을 찾을 수 없습니다")
                self.page.keyboard.press('Escape')
                return {u: False for u in usernames}

            # 각 사용자 검색
            print(f"\n[검색] 팔로워 확인 중... (제한 방지를 위해 천천히 진행)")
            follower_count = 0
            non_follower_count = 0

            for i, username in enumerate(usernames, 1):
                if username in self.follower_cache:
                    results[username] = self.follower_cache[username]
                    if results[username]:
                        follower_count += 1
                    else:
                        non_follower_count += 1
                    continue

                try:
                    # 검색
                    search_input.fill('')
                    self._wait(0.5)
                    search_input.fill(username)
                    # 3~5초 랜덤 대기 (제한 방지)
                    wait_time = random.uniform(3, 5)
                    self._wait(wait_time)

                    # 결과 확인
                    result_link = self.page.locator(f'a[href="/{username}/"] span:has-text("{username}")').first
                    is_follower = result_link.count() > 0

                    results[username] = is_follower
                    self.follower_cache[username] = is_follower

                    if is_follower:
                        follower_count += 1
                        print(f"  [{i}/{total}] {username}: ✅ 팔로워")
                    else:
                        non_follower_count += 1
                        print(f"  [{i}/{total}] {username}: ❌ 비팔로워")

                    # 10명마다 추가 대기 (15초)
                    if i % 10 == 0 and i < total:
                        print(f"  → 제한 방지를 위해 15초 대기 중...")
                        self._wait(15)

                except Exception as e:
                    results[username] = False
                    self.follower_cache[username] = False
                    non_follower_count += 1
                    print(f"  [{i}/{total}] {username}: ⏭️ 확인 실패")

            # 팔로워 모달 닫기
            self.page.keyboard.press('Escape')
            self._wait(0.5)

            print(f"\n  → 팔로워 확인 완료")
            print(f"  → 팔로워: {follower_count}명 / 비팔로워: {non_follower_count}명")

        except Exception as e:
            print(f"  ❌ 팔로우 확인 오류: {e}")
            for username in usernames:
                if username not in results:
                    results[username] = False

        return results

    def _convert_utc_to_kst(self, utc_str: str) -> str:
        """UTC 시간을 KST(한국 시간)로 변환"""
        if not utc_str:
            return ""
        try:
            # ISO 형식 파싱: 2026-01-14T10:10:38.000Z
            from datetime import datetime, timedelta
            utc_str = utc_str.replace('Z', '+00:00')
            dt = datetime.fromisoformat(utc_str.replace('.000', ''))
            # UTC+9 (한국 시간)
            kst = dt + timedelta(hours=9)
            return kst.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            return utc_str

    def _save_to_excel(self, comments: List[Dict], follow_status: Dict[str, bool]):
        """Excel 저장"""
        output_path = self.config_path.parent / self.config['output_file']

        print(f"\n{'='*80}")
        print("Excel 저장")
        print(f"{'='*80}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Instagram 댓글"

        # 헤더
        headers = ['번호', '닉네임', '댓글 내용', '작성시간', '답글 여부', '팔로우 여부']
        ws.append(headers)

        # 헤더 스타일
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # 데이터 추가
        for idx, comment in enumerate(comments, 1):
            username = comment['username']
            content = comment['content']
            datetime_kst = self._convert_utc_to_kst(comment.get('datetime', ''))
            is_reply = "[답글]" if comment.get('is_reply', False) else ""
            is_follower = "O" if follow_status.get(username, False) else "X"

            row = [idx, username, content, datetime_kst, is_reply, is_follower]
            ws.append(row)

            for col in range(1, len(row) + 1):
                ws.cell(row=idx + 1, column=col).border = thin_border

        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12

        # 저장
        wb.save(output_path)

        # 통계
        follower_count = sum(1 for v in follow_status.values() if v)
        reply_count = sum(1 for c in comments if c.get('is_reply', False))

        print(f"  💾 저장 완료: {output_path}")
        print(f"  → 총 댓글: {len(comments)}개")
        print(f"  → 답글: {reply_count}개")
        print(f"  → 팔로워: {follower_count}명 / {len(follow_status)}명")

    def run(self):
        """메인 실행"""
        print(f"\n{'='*80}")
        print("Instagram 댓글 크롤러")
        print(f"{'='*80}")
        print(f"  대상 게시물: {self.config['post_url']}")
        print(f"  게시물 작성자: {self.config['post_author']}")
        print(f"{'='*80}")

        try:
            # 1. 브라우저 시작
            self._start_browser()

            # 2. 로그인
            if not self.login():
                print(f"\n  ❌ 로그인 실패. 프로그램을 종료합니다.")
                return

            # 3. 스크롤 내리기 (숨겨진 댓글 보기 버튼까지)
            self._scroll_until_hidden_comments()

            # 4. 숨겨진 댓글 보기
            self._click_hidden_comments()

            # 5. 답글 펼치기
            self._click_reply_buttons()

            # 6. 댓글 추출
            comments = self._extract_comments()

            if not comments:
                print(f"\n  ❌ 댓글을 찾을 수 없습니다.")
                print(f"  → F12 개발자 도구로 HTML 구조를 확인해주세요.")
                input("\n계속하려면 Enter를 누르세요...")
                return

            # 7. 팔로우 여부 확인
            unique_usernames = list(set(c['username'] for c in comments))
            follow_status = self._check_follow_status(unique_usernames)

            # 8. Excel 저장
            self._save_to_excel(comments, follow_status)

            print(f"\n{'='*80}")
            print("✅ 크롤링 완료!")
            print(f"{'='*80}")

        except Exception as e:
            print(f"\n  ❌ 오류 발생: {e}")
            import traceback
            traceback.print_exc()

        finally:
            input("\n종료하려면 Enter를 누르세요...")
            self._close_browser()


def main():
    """메인 함수"""
    try:
        crawler = InstagramCrawler()
        crawler.run()
    except FileNotFoundError as e:
        print(f"❌ 오류: {e}")
        print("\nconfig.json 파일에 Instagram 계정 정보를 입력해주세요.")
    except Exception as e:
        print(f"❌ 오류: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
