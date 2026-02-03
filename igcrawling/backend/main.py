"""
Instagram 댓글 크롤러 API
FastAPI + Playwright
"""

import asyncio
import sys
import uuid
import io
import logging
from datetime import datetime
from typing import Dict

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Windows에서 asyncio subprocess 지원을 위한 설정
if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
# 개발 중 비활성화
# from slowapi import Limiter, _rate_limit_exceeded_handler
# from slowapi.util import get_remote_address
# from slowapi.errors import RateLimitExceeded
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from models import (
    CrawlRequest,
    CrawlProgress,
    CrawlResult,
    TaskStatus,
    CommentData
)
from crawler import run_crawler_async


# Rate limiter 설정 (개발 중 비활성화)
# limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="Instagram Comment Crawler API",
    description="Instagram 게시물 댓글 크롤링 서비스",
    version="1.0.0"
)

# Rate limit 에러 핸들러 (개발 중 비활성화)
# app.state.limiter = limiter
# app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# CORS 설정 (Vercel 도메인 허용)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "https://*.vercel.app",
        "*"  # 개발 중에는 모든 origin 허용 (배포 시 제한)
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 작업 상태 저장소
tasks: Dict[str, CrawlProgress] = {}
results: Dict[str, CrawlResult] = {}


async def update_task_progress(progress: CrawlProgress):
    """작업 진행 상태 업데이트 콜백"""
    tasks[progress.task_id] = progress


async def run_crawler(task_id: str, request: CrawlRequest):
    """크롤러 실행 (백그라운드)"""
    logger.info(f"[CRAWLER] Starting crawler for task {task_id}")
    try:
        result = await run_crawler_async(
            post_url=request.post_url,
            post_author=request.post_author,
            instagram_id=request.instagram_id,
            instagram_password=request.instagram_password,
            check_followers=request.check_followers,
            progress_callback=update_task_progress,
            task_id=task_id
        )

        logger.info(f"[CRAWLER] Result for {task_id}: success={result.get('success')}, comments={len(result.get('comments', []))}")

        # 결과 저장
        if result["success"]:
            results[task_id] = CrawlResult(
                task_id=task_id,
                status=TaskStatus.COMPLETED,
                comments=result["comments"],
                total_comments=len(result["comments"]),
                follower_count=result["follower_count"],
                non_follower_count=result["non_follower_count"]
            )
            logger.info(f"[CRAWLER] Result saved for {task_id}")
        else:
            results[task_id] = CrawlResult(
                task_id=task_id,
                status=TaskStatus.FAILED,
                error=result.get("error", "Unknown error")
            )
            logger.error(f"[CRAWLER] Task {task_id} failed: {result.get('error')}")
    except Exception as e:
        logger.exception(f"[CRAWLER] Exception for {task_id}: {e}")
        results[task_id] = CrawlResult(
            task_id=task_id,
            status=TaskStatus.FAILED,
            error=str(e)
        )


@app.get("/")
async def root():
    """API 상태 확인"""
    return {
        "status": "ok",
        "message": "Instagram Comment Crawler API",
        "version": "1.0.0"
    }


@app.post("/api/crawl")
# @limiter.limit("100/hour")  # 개발 중에는 비활성화
async def start_crawl(request: Request, crawl_request: CrawlRequest):
    """
    크롤링 시작 (비동기)

    - Rate limit: 시간당 5회
    - 반환: task_id (상태 조회용)
    """
    task_id = str(uuid.uuid4())

    # 초기 상태 설정
    tasks[task_id] = CrawlProgress(
        task_id=task_id,
        status=TaskStatus.PENDING,
        message="크롤링 대기 중...",
        progress=0
    )

    # 백그라운드 작업 시작
    asyncio.create_task(run_crawler(task_id, crawl_request))

    return {
        "task_id": task_id,
        "message": "크롤링이 시작되었습니다."
    }


@app.get("/api/status/{task_id}")
async def get_status(task_id: str):
    """
    크롤링 진행 상태 조회

    - 폴링용 엔드포인트
    - 진행률, 현재 단계, 메시지 반환
    """
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="Task not found")

    return tasks[task_id]


@app.get("/api/result/{task_id}")
async def get_result(task_id: str):
    """
    크롤링 결과 조회 (JSON)

    - 완료된 작업의 결과 반환
    - 댓글 목록, 통계 포함
    """
    if task_id not in results:
        # 아직 진행 중인지 확인
        if task_id in tasks:
            status = tasks[task_id]
            if status.status not in [TaskStatus.COMPLETED, TaskStatus.FAILED]:
                raise HTTPException(status_code=202, detail="Task still in progress")
        raise HTTPException(status_code=404, detail="Result not found")

    return results[task_id]


@app.get("/api/result/{task_id}/excel")
async def download_excel(task_id: str):
    """
    크롤링 결과 Excel 다운로드

    - 완료된 작업의 결과를 Excel 파일로 반환
    """
    if task_id not in results:
        raise HTTPException(status_code=404, detail="Result not found")

    result = results[task_id]

    if result.status != TaskStatus.COMPLETED:
        raise HTTPException(status_code=400, detail="Task not completed successfully")

    # Excel 파일 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "Instagram 댓글"

    # 헤더
    headers = ['번호', '닉네임', '댓글 내용', '작성시간', '답글 여부', '팔로우 여부']
    ws.append(headers)

    # 헤더 스타일
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 데이터 추가
    for idx, comment in enumerate(result.comments, 1):
        is_reply = "[답글]" if comment.is_reply else ""
        is_follower = "O" if comment.is_follower else "X"

        row = [
            idx,
            comment.username,
            comment.content,
            comment.datetime or "",
            is_reply,
            is_follower
        ]
        ws.append(row)

        for col in range(1, len(row) + 1):
            ws.cell(row=idx + 1, column=col).border = thin_border

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # 메모리에 저장
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 파일명 생성
    filename = f"instagram_comments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@app.delete("/api/task/{task_id}")
async def delete_task(task_id: str):
    """
    작업 삭제

    - 완료된 작업 정리용
    """
    deleted = False

    if task_id in tasks:
        del tasks[task_id]
        deleted = True

    if task_id in results:
        del results[task_id]
        deleted = True

    if not deleted:
        raise HTTPException(status_code=404, detail="Task not found")

    return {"message": "Task deleted successfully"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
